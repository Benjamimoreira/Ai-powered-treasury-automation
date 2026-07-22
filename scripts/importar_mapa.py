"""Script de migração único: lê as linhas de um dia do Mapa de Pagamentos
e Recebimentos (a folha com o nome do dia, ex. '21') e importa-as para a
tabela linhas_mapa. Só lê o .xlsx - nunca escreve nele.

Uso:
    python scripts/importar_mapa.py <caminho_mapa.xlsx> <DD-MM-YYYY>

Convenção de sinal: os valores de pagamentos ficam negativos em
linhas_mapa (mesmo sinal que um débito no extrato bancário /
MovimentoBancario.valor), mesmo que na folha do Mapa apareçam sempre
positivos. Os recebimentos mantêm o sinal positivo da folha. Isto permite
a reconciliar_dia comparar previsto/pago diretamente com o valor bruto do
extrato, sem ter de tratar pagamento e recebimento como casos à parte.
"""
import sys
from datetime import datetime
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import openpyxl

from app.db.models import LinhaMapa
from app.db.session import SessionLocal

COL_RECEBIMENTOS = dict(desc="B", previsto="C", real="D", empresa="E", imputacao="F")
COL_PAGAMENTOS = dict(desc="H", previsto="I", real="J", empresa="K", imputacao="L")


def encontrar_linha_totais(ws):
    for row in ws.iter_rows(min_row=5):
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("=SUM("):
                return cell.row
    raise RuntimeError("Não encontrei a linha de totais (fórmula =SUM) na folha.")


def importar_linhas(db, ws, dia, cols, tipo, linha_totais, sinal):
    total = 0
    for r in range(5, linha_totais):
        empresa = ws[f"{cols['empresa']}{r}"].value
        if not empresa:
            continue
        previsto = ws[f"{cols['previsto']}{r}"].value
        pago = ws[f"{cols['real']}{r}"].value
        imputacao = ws[f"{cols['imputacao']}{r}"].value
        db.add(LinhaMapa(
            dia=dia,
            tipo=tipo,
            linha=r,
            empresa=str(empresa).strip(),
            previsto=sinal * previsto if isinstance(previsto, (int, float)) else None,
            pago=sinal * pago if isinstance(pago, (int, float)) else None,
            imputacao=str(imputacao).strip() if imputacao else None,
        ))
        total += 1
    return total


def main():
    if len(sys.argv) != 3:
        print("Uso: python scripts/importar_mapa.py <caminho_mapa.xlsx> <DD-MM-YYYY>")
        sys.exit(1)

    caminho, data_str = sys.argv[1], sys.argv[2]
    dia = datetime.strptime(data_str, "%d-%m-%Y").date()
    nome_folha = f"{dia.day:02d}"

    wb = openpyxl.load_workbook(caminho, data_only=False)
    if nome_folha not in wb.sheetnames:
        print(f"A folha '{nome_folha}' não existe em {caminho}")
        sys.exit(1)
    ws = wb[nome_folha]
    linha_totais = encontrar_linha_totais(ws)

    db = SessionLocal()
    try:
        n_receb = importar_linhas(db, ws, dia, COL_RECEBIMENTOS, "recebimento", linha_totais, sinal=1)
        n_pag = importar_linhas(db, ws, dia, COL_PAGAMENTOS, "pagamento", linha_totais, sinal=-1)
        db.commit()
    finally:
        db.close()

    print(f"Importadas {n_receb} linhas de recebimento e {n_pag} linhas de pagamento para o dia {data_str}.")


if __name__ == "__main__":
    main()
