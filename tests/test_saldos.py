from datetime import date

import openpyxl

from app.db.models import SaldoDiario
from app.services.saldos import listar_saldos_atuais, registar_saldos_do_dia, saldo_total_geral

DIA = date(2026, 7, 21)


def _criar_extrato_falso(pasta, nome_ficheiro, contabilistico, disponivel):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "Saldo contabilístico"
    ws["B1"] = f"{contabilistico} EUR"
    ws["A2"] = "Saldo disponível"
    ws["B2"] = f"{disponivel} EUR"
    caminho = pasta / nome_ficheiro
    wb.save(caminho)
    return caminho


def test_registar_saldos_do_dia_le_extrato_e_grava_na_bd(db_session, tmp_path):
    _criar_extrato_falso(tmp_path, "21-07-2026_Empresa Teste,LDA.xlsx", "1.234,56", "1.200,00")

    total = registar_saldos_do_dia(db_session, DIA, str(tmp_path))

    assert total == 1
    saldo = db_session.query(SaldoDiario).one()
    assert saldo.entidade == "Empresa Teste,LDA"
    assert saldo.saldo_contabilistico == 1234.56
    assert saldo.saldo_disponivel == 1200.0


def test_registar_saldos_do_dia_e_idempotente(db_session, tmp_path):
    _criar_extrato_falso(tmp_path, "21-07-2026_Empresa Teste,LDA.xlsx", "1.234,56", "1.200,00")

    primeira = registar_saldos_do_dia(db_session, DIA, str(tmp_path))
    segunda = registar_saldos_do_dia(db_session, DIA, str(tmp_path))

    # chamar duas vezes para o mesmo dia (ex.: sync automático + endpoint
    # manual /saldos/atualizar/{dia}) não pode duplicar a entidade - bug
    # real encontrado em produção (30 entidades duplicadas no dia 23/07,
    # distorcendo a média móvel da previsão de saldos).
    assert primeira == 1
    assert segunda == 0
    assert db_session.query(SaldoDiario).count() == 1


def test_saldo_total_geral_soma_o_ultimo_saldo_de_cada_entidade(db_session):
    db_session.add(SaldoDiario(
        dia=date(2026, 7, 20), entidade="Empresa A", saldo_contabilistico=100.0, saldo_disponivel=90.0,
    ))
    db_session.add(SaldoDiario(
        dia=date(2026, 7, 21), entidade="Empresa A", saldo_contabilistico=150.0, saldo_disponivel=140.0,
    ))
    db_session.add(SaldoDiario(
        dia=date(2026, 7, 20), entidade="Empresa B", saldo_contabilistico=50.0, saldo_disponivel=50.0,
    ))
    db_session.commit()

    resultado = saldo_total_geral(db_session)

    # Empresa A conta só a leitura mais recente (150), não as duas (100+150)
    assert resultado == {
        "entidades": 2,
        "saldo_contabilistico_total": 200.0,
        "saldo_disponivel_total": 190.0,
    }


def test_listar_saldos_atuais_devolve_so_a_leitura_mais_recente(db_session):
    db_session.add(SaldoDiario(
        dia=date(2026, 7, 20), entidade="Empresa A", saldo_contabilistico=100.0, saldo_disponivel=90.0,
    ))
    db_session.add(SaldoDiario(
        dia=date(2026, 7, 21), entidade="Empresa A", saldo_contabilistico=150.0, saldo_disponivel=140.0,
    ))
    db_session.commit()

    resultado = listar_saldos_atuais(db_session)

    assert len(resultado) == 1
    assert resultado[0].saldo_contabilistico == 150.0


def test_saldo_total_geral_com_dia_usa_a_leitura_ate_esse_dia(db_session):
    db_session.add(SaldoDiario(
        dia=date(2026, 7, 20), entidade="Empresa A", saldo_contabilistico=100.0, saldo_disponivel=90.0,
    ))
    db_session.add(SaldoDiario(
        dia=date(2026, 7, 24), entidade="Empresa A", saldo_contabilistico=250000.0, saldo_disponivel=57000.0,
    ))
    db_session.commit()

    # Ao selecionar o dia 21 no dashboard, o total tem de mostrar o saldo
    # como estava nesse dia (100/90) - não o mais recente de sempre
    # (250000/57000, só disponível a partir do dia 24). Reproduz o bug em
    # que a Visão Geral mostrava sempre o mesmo total ao mudar o dia.
    resultado = saldo_total_geral(db_session, dia=date(2026, 7, 21))
    assert resultado == {
        "entidades": 1,
        "saldo_contabilistico_total": 100.0,
        "saldo_disponivel_total": 90.0,
    }

    resultado_depois = saldo_total_geral(db_session, dia=date(2026, 7, 25))
    assert resultado_depois["saldo_contabilistico_total"] == 250000.0


def test_saldo_total_geral_sem_dia_continua_a_usar_o_mais_recente(db_session):
    db_session.add(SaldoDiario(
        dia=date(2026, 7, 20), entidade="Empresa A", saldo_contabilistico=100.0, saldo_disponivel=90.0,
    ))
    db_session.add(SaldoDiario(
        dia=date(2026, 7, 24), entidade="Empresa A", saldo_contabilistico=250000.0, saldo_disponivel=57000.0,
    ))
    db_session.commit()

    resultado = saldo_total_geral(db_session)
    assert resultado["saldo_contabilistico_total"] == 250000.0
