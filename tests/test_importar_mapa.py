from datetime import date

import openpyxl

from app.db.models import LinhaMapa
from scripts.importar_mapa import COL_PAGAMENTOS, COL_RECEBIMENTOS, encontrar_linha_totais, importar_linhas

DIA = date(2026, 7, 21)


def _criar_folha_do_dia(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "21"
    ws["A4"] = "Nº"
    ws["E5"] = "Ancora Apogeu"
    ws["D5"] = 166.29
    ws["K5"] = "Ancora Apogeu"
    ws["I5"] = 45
    ws["K6"] = "Palavra"
    ws["J6"] = 359.96
    ws["A8"] = "=SUM(C5:C7)"
    caminho = tmp_path / "mapa.xlsx"
    wb.save(caminho)
    return caminho


def test_importar_linhas_aplica_sinal_correto_por_tipo(tmp_path, db_session):
    caminho = _criar_folha_do_dia(tmp_path)
    wb = openpyxl.load_workbook(caminho, data_only=False)
    ws = wb["21"]
    linha_totais = encontrar_linha_totais(ws)

    n_receb = importar_linhas(db_session, ws, DIA, COL_RECEBIMENTOS, "recebimento", linha_totais, sinal=1)
    n_pag = importar_linhas(db_session, ws, DIA, COL_PAGAMENTOS, "pagamento", linha_totais, sinal=-1)
    db_session.commit()

    assert n_receb == 1
    assert n_pag == 2

    receb = db_session.query(LinhaMapa).filter_by(tipo="recebimento").one()
    assert receb.pago == 166.29

    pag_previsto = db_session.query(LinhaMapa).filter_by(tipo="pagamento", empresa="Ancora Apogeu").one()
    assert pag_previsto.previsto == -45

    pag_pago = db_session.query(LinhaMapa).filter_by(tipo="pagamento", empresa="Palavra").one()
    assert pag_pago.pago == -359.96
