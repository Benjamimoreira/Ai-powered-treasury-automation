import os
from datetime import date

import openpyxl
import pytest

from app.db.models import LinhaMapa, MovimentoBancario, SaldoDiario
from app.services import onedrive_sync

DIA = date(2026, 7, 21)


def _criar_extrato_falso(pasta, nome_ficheiro):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "Saldo contabilístico"
    ws["B1"] = "100,00 EUR"
    ws["A2"] = "Saldo disponível"
    ws["B2"] = "90,00 EUR"
    ws["A5"] = "Data mov."
    ws["A6"] = "21-07-2026"
    ws["B6"] = "21-07-2026"
    ws["C6"] = "TRANSF TESTE"
    ws["D6"] = "-50,00"
    caminho = os.path.join(pasta, nome_ficheiro)
    wb.save(caminho)
    return caminho


def _preparar_pasta_extratos(tmp_path):
    pasta = tmp_path / "FINANCEIRO" / "03 - Extratos Bancários" / "Movimentos Diários" / "CGD" / "07_Julho" / "21-07-2026"
    pasta.mkdir(parents=True)
    _criar_extrato_falso(pasta, "21-07-2026_Empresa Teste,LDA.xlsx")
    return pasta


def test_atualizar_dados_recentes_falha_sem_onedrive_raiz(db_session, monkeypatch):
    monkeypatch.delenv("ONEDRIVE_RAIZ", raising=False)
    with pytest.raises(RuntimeError):
        onedrive_sync.atualizar_dados_recentes(db_session, dias_atras=0)


def test_atualizar_dados_recentes_importa_movimentos_e_saldos_novos(db_session, monkeypatch, tmp_path):
    _preparar_pasta_extratos(tmp_path)
    monkeypatch.setenv("ONEDRIVE_RAIZ", str(tmp_path))
    monkeypatch.setattr(onedrive_sync, "date", _DataFixa)

    resultado = onedrive_sync.atualizar_dados_recentes(db_session, dias_atras=0)

    assert resultado["dias_com_movimentos_novos"] == ["2026-07-21"]
    assert resultado["dias_com_saldos_novos"] == ["2026-07-21"]
    assert resultado["erros"] == []
    assert db_session.query(MovimentoBancario).count() == 1
    assert db_session.query(SaldoDiario).count() == 1


def test_atualizar_dados_recentes_e_idempotente(db_session, monkeypatch, tmp_path):
    _preparar_pasta_extratos(tmp_path)
    monkeypatch.setenv("ONEDRIVE_RAIZ", str(tmp_path))
    monkeypatch.setattr(onedrive_sync, "date", _DataFixa)

    onedrive_sync.atualizar_dados_recentes(db_session, dias_atras=0)
    segunda = onedrive_sync.atualizar_dados_recentes(db_session, dias_atras=0)

    assert segunda["dias_com_movimentos_novos"] == []
    assert segunda["dias_com_saldos_novos"] == []
    assert db_session.query(MovimentoBancario).count() == 1


def test_atualizar_dados_recentes_reimporta_saldos_de_ontem(db_session, monkeypatch, tmp_path):
    """O CGD substitui a pasta do dia anterior pela versão final às 8:30 da
    manhã seguinte - um saldo já importado (provisório, ex. de uma corrida
    da tarde anterior) tem de ser substituído pelo valor atual do extrato,
    não ignorado como "já importado"."""
    _preparar_pasta_extratos(tmp_path)  # extrato de 21-07, saldo disponível = 90,00
    monkeypatch.setenv("ONEDRIVE_RAIZ", str(tmp_path))

    class _DataFixaDiaSeguinte(date):
        @classmethod
        def today(cls):
            return date(2026, 7, 22)

    monkeypatch.setattr(onedrive_sync, "date", _DataFixaDiaSeguinte)

    # saldo provisório já importado antes, com um valor diferente do que
    # está agora no extrato - simula a versão da tarde anterior.
    db_session.add(SaldoDiario(
        dia=DIA, entidade="Empresa Teste,LDA", saldo_contabilistico=50.0, saldo_disponivel=10.0,
    ))
    db_session.commit()

    resultado = onedrive_sync.atualizar_dados_recentes(db_session, dias_atras=1)

    assert resultado["dias_com_saldos_novos"] == ["2026-07-21"]
    saldo = db_session.query(SaldoDiario).filter(SaldoDiario.dia == DIA).one()
    assert saldo.saldo_disponivel == 90.0


def test_atualizar_dados_recentes_ignora_dia_sem_pasta(db_session, monkeypatch, tmp_path):
    monkeypatch.setenv("ONEDRIVE_RAIZ", str(tmp_path))
    monkeypatch.setattr(onedrive_sync, "date", _DataFixa)

    resultado = onedrive_sync.atualizar_dados_recentes(db_session, dias_atras=0)

    assert resultado == {
        "dias_verificados": 1,
        "dias_com_movimentos_novos": [],
        "dias_com_saldos_novos": [],
        "dias_com_mapa_novo": [],
        "erros": [],
    }


class _DataFixa(date):
    @classmethod
    def today(cls):
        return date(2026, 7, 21)
