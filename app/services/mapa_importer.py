"""Lógica de leitura das linhas do Mapa de Pagamentos e Recebimentos (a
folha com o nome do dia, ex. '21') - partilhada entre o script de
migração único (scripts/importar_mapa.py) e a sincronização automática a
partir do OneDrive (app/services/onedrive_sync.py). Só lê o .xlsx - nunca
escreve nele.

Convenção de sinal: os valores de pagamentos ficam negativos em
linhas_mapa (mesmo sinal que um débito no extrato bancário /
MovimentoBancario.valor), mesmo que na folha do Mapa apareçam sempre
positivos. Os recebimentos mantêm o sinal positivo da folha. Isto permite
a reconciliar_dia comparar previsto/pago diretamente com o valor bruto do
extrato, sem ter de tratar pagamento e recebimento como casos à parte.
"""
from app.db.models import LinhaMapa

COL_RECEBIMENTOS = dict(desc="B", previsto="C", real="D", empresa="E", imputacao="F")
COL_PAGAMENTOS = dict(desc="H", previsto="I", real="J", empresa="K", imputacao="L")


def encontrar_linha_totais(ws):
    for row in ws.iter_rows(min_row=5):
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("=SUM("):
                return cell.row
    raise RuntimeError("Não encontrei a linha de totais (fórmula =SUM) na folha.")


def importar_linhas(db, ws, dia, cols, tipo, linha_totais, sinal):
    total = 0
    for r in range(5, linha_totais):
        empresa = ws[f"{cols['empresa']}{r}"].value
        if not empresa:
            continue
        previsto = ws[f"{cols['previsto']}{r}"].value
        pago = ws[f"{cols['real']}{r}"].value
        imputacao = ws[f"{cols['imputacao']}{r}"].value
        db.add(LinhaMapa(
            dia=dia,
            tipo=tipo,
            linha=r,
            empresa=str(empresa).strip(),
            previsto=sinal * previsto if isinstance(previsto, (int, float)) else None,
            pago=sinal * pago if isinstance(pago, (int, float)) else None,
            imputacao=str(imputacao).strip() if imputacao else None,
        ))
        total += 1
    return total


def importar_dia_do_mapa(db, caminho_mapa: str, dia) -> tuple:
    """Abre o ficheiro do Mapa, localiza a folha do dia e importa as
    linhas de recebimento e pagamento. Devolve (n_recebimentos,
    n_pagamentos). Levanta FileNotFoundError/KeyError se a folha do dia
    não existir."""
    import openpyxl

    nome_folha = f"{dia.day:02d}"
    wb = openpyxl.load_workbook(caminho_mapa, data_only=False)
    if nome_folha not in wb.sheetnames:
        raise KeyError(f"A folha '{nome_folha}' não existe em {caminho_mapa}")
    ws = wb[nome_folha]
    linha_totais = encontrar_linha_totais(ws)

    n_receb = importar_linhas(db, ws, dia, COL_RECEBIMENTOS, "recebimento", linha_totais, sinal=1)
    n_pag = importar_linhas(db, ws, dia, COL_PAGAMENTOS, "pagamento", linha_totais, sinal=-1)
    return n_receb, n_pag
