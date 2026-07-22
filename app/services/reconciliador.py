import os
import re
import shutil
import tempfile
import time
import unicodedata

import openpyxl
from sqlalchemy.orm import Session

from app.db.models import CasoAmbiguo, LinhaMapa, MovimentoBancario, Reconciliacao

TOLERANCIA_VALOR = 0.01

PALAVRAS_IGNORAR = {"LDA", "SA", "DE", "DA", "DO", "E"}


def abrir_workbook_com_retry(caminho, tentativas=5, espera=1.0):
    """Abre um .xlsx com openpyxl, tentando novamente em caso de
    PermissionError transitório (ficheiros sincronizados no OneDrive,
    antivírus a indexar, ficheiro aberto no Excel). Lê a partir de uma
    cópia temporária para não depender de o ficheiro original ficar
    livre a meio da tentativa."""
    ultimo_erro = None
    for tentativa in range(tentativas):
        try:
            with tempfile.TemporaryDirectory() as tmp:
                destino = os.path.join(tmp, os.path.basename(caminho))
                shutil.copyfile(caminho, destino)
                return openpyxl.load_workbook(destino, data_only=True)
        except PermissionError as e:
            ultimo_erro = e
            if tentativa < tentativas - 1:
                time.sleep(espera)
    raise PermissionError(
        f"Não consegui ler '{caminho}' depois de {tentativas} tentativas "
        f"(ficheiro continua bloqueado)."
    ) from ultimo_erro


def ler_movimentos_do_extrato(caminho):
    wb = abrir_workbook_com_retry(caminho)
    ws = wb.active

    linha_cabecalho = None
    for row in ws.iter_rows(min_row=1, max_row=20):
        if row[0].value and str(row[0].value).strip().startswith("Data mov"):
            linha_cabecalho = row[0].row
            break
    if linha_cabecalho is None:
        return []

    movimentos = []
    for row in ws.iter_rows(min_row=linha_cabecalho + 1):
        data_mov, _data_valor, descricao, montante = row[0].value, row[1].value, row[2].value, row[3].value
        if data_mov is None and descricao is None:
            break
        if montante is None:
            continue
        valor = float(str(montante).strip().replace(".", "").replace(",", "."))
        movimentos.append({
            "descricao": str(descricao).strip() if descricao else "",
            "valor": valor,
        })
    return movimentos


def remover_acentos(texto):
    nfkd = unicodedata.normalize("NFKD", texto)
    return "".join(c for c in nfkd if not unicodedata.combining(c))


def normalizar(texto):
    texto = remover_acentos(str(texto)).upper()
    return re.sub(r"[^A-Z0-9]", "", texto)


def nome_empresa_do_ficheiro(caminho):
    nome = os.path.splitext(os.path.basename(caminho))[0]
    nome = re.sub(r"^\d{2}-\d{2}-\d{4}_", "", nome)
    return nome.strip()


def chave_empresa(nome_completo):
    """Chave normalizada do nome da empresa sem palavras tipo LDA/SA, para
    que 'ANCORA APOGEU,LDA' e 'Ancora Apogeu' (sem sigla) batam certo."""
    partes = [p for p in re.split(r"[\s,]+", str(nome_completo).strip()) if p]
    significativas = [p for p in partes if normalizar(p) not in PALAVRAS_IGNORAR]
    return normalizar(" ".join(significativas))


def importar_extrato_para_bd(db: Session, caminho: str, dia, empresa: str) -> int:
    """Lê um ficheiro de extrato (.xlsx) e grava cada movimento como uma
    linha em movimentos_bancarios. Devolve o número de movimentos
    inseridos. Não faz deteção de duplicados (isso é Fase 2 -
    reconciliação); esta função só migra os dados brutos do Excel para
    a base de dados."""
    movimentos = ler_movimentos_do_extrato(caminho)
    for mov in movimentos:
        db.add(MovimentoBancario(
            dia=dia,
            empresa=empresa,
            descricao=mov["descricao"],
            valor=mov["valor"],
            ficheiro_origem=os.path.basename(caminho),
        ))
    db.commit()
    return len(movimentos)


def _linha_bate_com_movimento(linha: LinhaMapa, movimento: MovimentoBancario) -> bool:
    if linha.previsto is None or linha.pago is not None:
        return False
    return (
        chave_empresa(linha.empresa) == chave_empresa(movimento.empresa)
        and abs(linha.previsto - movimento.valor) < TOLERANCIA_VALOR
    )


def reconciliar_dia(db: Session, dia) -> dict:
    """Corre a reconciliação de um dia: tenta casar cada movimento bancário
    ainda não processado com uma linha do mapa (linhas_mapa) do mesmo dia,
    pela empresa (chave_empresa) e pelo valor (com tolerância). Grava o
    resultado em Reconciliacao (e CasoAmbiguo quando há mais que uma linha
    candidata). Correr duas vezes para o mesmo dia não duplica trabalho -
    só processa movimentos que ainda não têm nenhuma Reconciliacao."""
    movimentos = (
        db.query(MovimentoBancario)
        .filter(MovimentoBancario.dia == dia)
        .filter(~MovimentoBancario.reconciliacoes.any())
        .all()
    )
    linhas = db.query(LinhaMapa).filter(LinhaMapa.dia == dia).all()

    casados = novos = ambiguos = 0
    linhas_ja_usadas = set()

    for movimento in movimentos:
        candidatos = [
            linha for linha in linhas
            if linha.id not in linhas_ja_usadas and _linha_bate_com_movimento(linha, movimento)
        ]

        if len(candidatos) == 1:
            linha = candidatos[0]
            linha.pago = movimento.valor
            linhas_ja_usadas.add(linha.id)
            db.add(Reconciliacao(movimento_id=movimento.id, linha_id=linha.id, tipo_match="exato"))
            casados += 1
        elif len(candidatos) > 1:
            db.add(Reconciliacao(movimento_id=movimento.id, linha_id=None, tipo_match="ambiguo"))
            db.add(CasoAmbiguo(
                movimento_id=movimento.id,
                dia=dia,
                empresa=movimento.empresa,
                valor=movimento.valor,
                candidatos=[linha.id for linha in candidatos],
            ))
            ambiguos += 1
        else:
            db.add(Reconciliacao(movimento_id=movimento.id, linha_id=None, tipo_match="novo"))
            novos += 1

    db.commit()
    return {"casados": casados, "novos": novos, "ambiguos": ambiguos}


def auditoria_dia(db: Session, dia) -> dict:
    """Verificação read-only (não grava nada): conta movimentos sem
    correspondência numa linha do mapa (sem_match_fwd) e linhas do mapa
    com previsto ainda em aberto (sem pago) que não correspondem a nenhum
    movimento (sem_match_rev). Linhas com previsto E pago já preenchidos
    (resolvidas antes de existir esta API) ficam de fora - já não são
    "previstos por bater", são histórico. Independente de reconciliar_dia
    já ter corrido."""
    movimentos = db.query(MovimentoBancario).filter(MovimentoBancario.dia == dia).all()
    linhas = db.query(LinhaMapa).filter(
        LinhaMapa.dia == dia, LinhaMapa.previsto.isnot(None), LinhaMapa.pago.is_(None),
    ).all()

    sem_match_fwd = sum(
        1 for movimento in movimentos
        if not any(_linha_bate_com_movimento(linha, movimento) for linha in linhas)
    )
    sem_match_rev = sum(
        1 for linha in linhas
        if not any(_linha_bate_com_movimento(linha, movimento) for movimento in movimentos)
    )
    return {"sem_match_fwd": sem_match_fwd, "sem_match_rev": sem_match_rev}


def resolver_ambiguo(db: Session, caso_id: int, linha_id, resolvido_por: str) -> CasoAmbiguo:
    """Regista a decisão humana sobre um caso ambíguo: associa o movimento
    à linha escolhida (ou marca como 'novo' se linha_id for None)."""
    caso = db.get(CasoAmbiguo, caso_id)
    if caso is None:
        raise ValueError(f"Caso ambíguo {caso_id} não encontrado.")

    reconciliacao = (
        db.query(Reconciliacao)
        .filter(Reconciliacao.movimento_id == caso.movimento_id, Reconciliacao.tipo_match == "ambiguo")
        .first()
    )

    if linha_id is not None:
        linha = db.get(LinhaMapa, linha_id)
        if linha is None:
            raise ValueError(f"Linha {linha_id} não encontrada.")
        linha.pago = caso.valor
        caso.resolucao = f"linha_id={linha_id}"
        if reconciliacao is not None:
            reconciliacao.linha_id = linha_id
            reconciliacao.tipo_match = "exato"
    else:
        caso.resolucao = "novo"
        if reconciliacao is not None:
            reconciliacao.tipo_match = "novo"

    caso.resolvido_por = resolvido_por
    db.commit()
    return caso
