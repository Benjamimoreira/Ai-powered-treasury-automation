"""
Extrai dados de faturas (PDF digitalizado) da pasta "01 - Compras" do OneDrive
Contabilidade, para as empresas indicadas, percorrendo todos os meses.

Usa OCR (Tesseract) porque os PDFs sao digitalizacoes sem texto pesquisavel.
Escreve resultados incrementalmente num CSV (para nao perder trabalho se
parar a meio) e no fim gera um Excel (.xlsx) com uma folha por empresa.

Versão Docker/Linux de Ambiente de Trabalho/Fornecedores/extrair_faturas.py
- mesma lógica, só muda a origem dos dados (OneDrive montado em /onedrive) e
o destino da saída (pasta /output, volume dedicado); tesseract/poppler vêm
do apt (tesseract-ocr, tesseract-ocr-por, poppler-utils), já no PATH do
container, por isso sem caminhos fixos como na versão Windows.
"""

import csv
import os
import re
import unicodedata
from pathlib import Path

import pytesseract
from pdf2image import convert_from_path

ONEDRIVE_RAIZ = os.environ.get("ONEDRIVE_RAIZ", "/onedrive")
ANO = os.environ.get("FATURAS_OCR_ANO") or str(__import__("datetime").date.today().year)
RAIZ_ANO = os.path.join(ONEDRIVE_RAIZ, "CONTABILIDADE", "04 - Arquivos", ANO)

EMPRESAS = [
    "Palavradicional",
    "Habiserve Investimentos Imobiliários",
    "Habiserve Construções Norte",
]

PASTA_SAIDA = Path("/output")
PASTA_SAIDA.mkdir(exist_ok=True)
CSV_SAIDA = PASTA_SAIDA / "faturas_compras.csv"
XLSX_SAIDA = PASTA_SAIDA / "faturas_compras.xlsx"
PASTA_CACHE_OCR = PASTA_SAIDA / "ocr_cache"
PASTA_CACHE_OCR.mkdir(exist_ok=True)

CAMPOS = [
    "Empresa",
    "Mes",
    "Ficheiro",
    "Remetente",
    "Destinatario",
    "Email",
    "Numero Fatura",
    "Valor (EUR)",
    "Caminho completo",
]

RE_EMAIL = re.compile(r"[\w.+-]+@[\w-]+\.[\w.-]+")

# Palavras-chave que antecedem o numero da fatura, por ordem de prioridade.
RE_NUM_FATURA_KEYWORDS = re.compile(
    r"(?:Invoice\s*no\.?|N[ºo°.]*\s*da\s*fatura|Fatura[\s\-]*Recibo|"
    r"FACTURA\s*N[ºo°.]*|FATURA\s*N[ºo°.]*|N[ºo°.]*\s*Fatura)"
    r"\s*[:\-]?\s*[<]?\s*([A-Z0-9][A-Z0-9/\-\s]{2,20}[A-Z0-9])",
    re.IGNORECASE,
)
RE_NUM_FATURA_GENERICO = re.compile(r"\b(F[TRASC]\s?[A-Z]{0,3}\d[\dA-Z]{2,}\s?/\s?\d+)\b")
RE_ATCUD = re.compile(r"ATCUD\s*[:\-]?\s*([A-Z0-9\-]{6,})", re.IGNORECASE)

LINHA_IGNORAR_VALOR = re.compile(
    r"capital|iban|nif|nipc|contribuinte|telefone|telem[oó]vel|atcud|"
    r"vat\s*no|order\s*no|cap\.?\s*soc",
    re.IGNORECASE,
)

PALAVRAS_TOTAL = re.compile(
    r"total\s*a\s*pagar|tot\.?\s*c\s*/\s*iva|total\s*c\s*/\s*iva|valor\s*total|"
    r"total\s*da\s*fatura|total\s*do\s*documento|grand\s*total|total\s*paid|"
    r"amount\s*due|valor\s*a\s*pagar|^\s*total\s*:?\s*$",
    re.IGNORECASE | re.MULTILINE,
)
_NUM = r"\d{1,3}(?:[.\s]\d{3})*[,.]\d{2}"
RE_VALOR_NUM = re.compile(rf"(?:€\s*({_NUM}))|(?:({_NUM})\s*(?:EUR|€))")


def _valores_na_linha(linha: str) -> list:
    return [a or b for a, b in RE_VALOR_NUM.findall(linha)]

LINHA_IGNORAR_REMETENTE = re.compile(
    r"^\s*$|atcud|original|p[aá]g\.?\s*\d|processado por programa|"
    r"^invoice for|^account\s*id|^invoice\s*/\s*payment",
    re.IGNORECASE,
)


def limpar_valor(txt: str) -> str:
    v = txt.strip().replace(" ", "")
    # normaliza milhar/decimal: assume ultimo separador antes de 2 digitos é decimal
    if "," in v and "." in v:
        if v.rfind(",") > v.rfind("."):
            v = v.replace(".", "").replace(",", ".")
        else:
            v = v.replace(",", "")
    elif "," in v:
        v = v.replace(".", "").replace(",", ".")
    return v


def extrair_numero_fatura(texto: str) -> str:
    m = RE_NUM_FATURA_KEYWORDS.search(texto)
    if m:
        return re.sub(r"\s+", " ", m.group(1).strip())
    m = RE_NUM_FATURA_GENERICO.search(texto)
    if m:
        return m.group(1).strip()
    m = RE_ATCUD.search(texto)
    if m:
        return m.group(1).strip()
    return ""


def extrair_valor(texto: str) -> str:
    linhas_validas = [l for l in texto.splitlines() if not LINHA_IGNORAR_VALOR.search(l)]
    texto_filtrado = "\n".join(linhas_validas)

    # 1) procura um valor na mesma linha (ou nas duas seguintes) de uma palavra-chave de total
    linhas = texto_filtrado.splitlines()
    for i, linha in enumerate(linhas):
        if PALAVRAS_TOTAL.search(linha):
            candidatos = []
            for candidata in linhas[i : i + 3]:
                candidatos.extend(_valores_na_linha(candidata))
            if candidatos:
                return limpar_valor(candidatos[-1])

    # 2) fallback: maior valor monetario encontrado no texto filtrado
    valores = _valores_na_linha(texto_filtrado)
    if valores:
        try:
            return limpar_valor(max(valores, key=lambda v: float(limpar_valor(v))))
        except ValueError:
            return limpar_valor(valores[-1])
    return ""


def extrair_remetente(texto: str) -> str:
    m = re.search(r"^\s*NOME\s+(.+)$", texto, re.MULTILINE | re.IGNORECASE)
    if m:
        return m.group(1).strip()

    for linha in texto.splitlines():
        linha = linha.strip()
        if LINHA_IGNORAR_REMETENTE.search(linha):
            continue
        letras = sum(c.isalpha() for c in linha)
        if len(linha) >= 5 and letras / len(linha) >= 0.6:
            return linha
    return ""


def ocr_pdf(caminho: Path) -> str:
    # sem poppler_path/tesseract_cmd explícitos - tesseract-ocr e poppler-utils
    # (apt) já ficam no PATH do container
    paginas = convert_from_path(str(caminho), dpi=300)
    textos = [pytesseract.image_to_string(p, lang="por") for p in paginas]
    return "\n".join(textos)


def ocr_pdf_cache(caminho: Path, empresa: str, mes: str) -> str:
    pasta = PASTA_CACHE_OCR / empresa / mes
    pasta.mkdir(parents=True, exist_ok=True)
    ficheiro_cache = pasta / f"{caminho.stem}.txt"
    if ficheiro_cache.exists():
        return ficheiro_cache.read_text(encoding="utf-8")
    texto = ocr_pdf(caminho)
    ficheiro_cache.write_text(texto, encoding="utf-8")
    return texto


def main():
    with open(CSV_SAIDA, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=CAMPOS)
        writer.writeheader()

        for empresa in EMPRESAS:
            pasta_compras = Path(RAIZ_ANO) / empresa / "01 - Compras"
            if not pasta_compras.exists():
                print(f"[AVISO] pasta nao encontrada para {empresa}: {pasta_compras}")
                continue

            meses = sorted([p for p in pasta_compras.iterdir() if p.is_dir()])
            for mes_dir in meses:
                pdfs = sorted(mes_dir.glob("*.pdf"))
                for pdf in pdfs:
                    print(f"{empresa} / {mes_dir.name} / {pdf.name}")
                    try:
                        texto = ocr_pdf_cache(pdf, empresa, mes_dir.name)
                    except Exception as exc:
                        print(f"  ERRO ao processar {pdf}: {exc}")
                        continue

                    email_m = RE_EMAIL.search(texto)
                    linha = {
                        "Empresa": empresa,
                        "Mes": mes_dir.name,
                        "Ficheiro": pdf.name,
                        "Remetente": extrair_remetente(texto),
                        "Destinatario": empresa,
                        "Email": email_m.group(0) if email_m else "",
                        "Numero Fatura": extrair_numero_fatura(texto),
                        "Valor (EUR)": extrair_valor(texto),
                        "Caminho completo": str(pdf),
                    }
                    writer.writerow(linha)
                    f.flush()

    gerar_excel()
    print("Concluido.")


def gerar_excel():
    import openpyxl
    from openpyxl.utils import get_column_letter

    linhas_por_empresa = {}
    with open(CSV_SAIDA, newline="", encoding="utf-8") as f:
        for linha in csv.DictReader(f):
            linhas_por_empresa.setdefault(linha["Empresa"], []).append(linha)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for empresa, linhas in linhas_por_empresa.items():
        nome_folha = unicodedata.normalize("NFKD", empresa).encode("ascii", "ignore").decode()
        nome_folha = re.sub(r"[\\/*?:\[\]]", "", nome_folha)[:31] or "Empresa"
        ws = wb.create_sheet(title=nome_folha)
        ws.append(CAMPOS)
        for linha in linhas:
            ws.append([linha[c] for c in CAMPOS])
        for i, campo in enumerate(CAMPOS, start=1):
            largura = max(12, min(40, max(len(str(l[campo])) for l in linhas) + 2))
            ws.column_dimensions[get_column_letter(i)].width = largura
    wb.save(XLSX_SAIDA)
    print(f"Excel gerado em: {XLSX_SAIDA}")


if __name__ == "__main__":
    from monitorizacao_client import monitorizar
    with monitorizar("extrair_faturas"):
        main()
